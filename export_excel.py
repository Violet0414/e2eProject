#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导出测试用例到Excel"""

import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 读取测试用例文件
with open('/home/mxr/e2eProject/output/测试用例.md', 'r', encoding='utf-8') as f:
    content = f.read()

# 解析测试用例
test_cases = []
# 匹配用例块：### TC-CASE-XXX 到下一个 ### 或文件结束
case_pattern = re.compile(r'### (TC-CASE-\d+)：(.*?)(?=\n---\n### |$)', re.DOTALL)

# 先按模块分割
sections = re.split(r'\n---\n\n### ', content)
# 第一个section包含开头的模块统计部分，需要跳过
cases_in_sections = []

for i, section in enumerate(sections):
    if i == 0:
        # 检查是否包含TC-CASE
        if 'TC-CASE' in section:
            cases_in_sections.append(section)
    else:
        cases_in_sections.append('### ' + section)

# 解析每个用例
for section in cases_in_sections:
    # 提取用例ID和标题行
    header_match = re.match(r'### (TC-CASE-\d+)：(.*?)\n', section)
    if not header_match:
        continue

    case_id = header_match.group(1)
    title = header_match.group(2).strip()

    # 提取关联测试点
    tp_match = re.search(r'\*\*关联测试点\*\*：(.+)', section)
    tp = tp_match.group(1).strip() if tp_match else ''

    # 提取优先级
    pri_match = re.search(r'\*\*优先级\*\*：(.+)', section)
    priority = pri_match.group(1).strip() if pri_match else ''

    # 提取前置条件
    pre_match = re.search(r'\*\*前置条件\*\*：(.*?)(?=\n\| 步骤 |---)', section, re.DOTALL)
    precondition = ''
    if pre_match:
        pre_text = pre_match.group(1).strip()
        # 清理换行
        precondition = ' '.join(pre_text.split())

    # 提取所属模块 - 根据用例ID判断
    module = ''
    if tp.startswith('TP01') or tp.startswith('TP02') or tp.startswith('TP03'):
        module = '文字统一'
    elif tp.startswith('TP04') or tp.startswith('TP05') or tp.startswith('TP06'):
        module = '被投诉单位信息'
    elif tp.startswith('TP07') or tp.startswith('TP08') or tp.startswith('TP09') or tp.startswith('TP10') or \
            tp.startswith('TP11') or tp.startswith('TP12') or tp.startswith('TP13') or tp.startswith('TP14') or \
            tp.startswith('TP15') or tp.startswith('TP16'):
        module = '涉及领域'
    elif tp.startswith('TP17') or tp.startswith('TP18') or tp.startswith('TP19') or \
            tp.startswith('TP20') or tp.startswith('TP21'):
        module = '案件来源'
    elif tp.startswith('TP22') or tp.startswith('TP23') or tp.startswith('TP24'):
        module = '涉及人数'
    elif tp.startswith('TP25') or tp.startswith('TP26') or tp.startswith('TP27'):
        module = '附件预览'
    elif tp.startswith('TP28') or tp.startswith('TP29') or tp.startswith('TP30'):
        module = '打印功能'
    elif tp.startswith('TP31'):
        module = '导出模板'
    elif tp.startswith('TP32') or tp.startswith('TP33') or tp.startswith('TP34') or tp.startswith('TP35'):
        module = '指派功能'
    elif tp.startswith('TP36') or tp.startswith('TP37') or tp.startswith('TP38'):
        module = '数据权限'
    elif tp.startswith('TP39') or tp.startswith('TP40'):
        module = '主办监察员'
    elif tp.startswith('TP41') or tp.startswith('TP42') or tp.startswith('TP43'):
        module = '受理状态'
    elif tp.startswith('TP44') or tp.startswith('TP45') or tp.startswith('TP46') or tp.startswith('TP47'):
        module = '结案相关'
    elif tp.startswith('TP48') or tp.startswith('TP49'):
        module = '统计报表'

    # 提取步骤表格
    steps = []
    table_match = re.search(r'\| 步骤 \| 操作 \| 测试数据 \| 预期结果 \|\n\|[^\n]+\n((?:\|[^\n]+\n)+)', section)
    if table_match:
        table_lines = table_match.group(1).strip().split('\n')
        for line in table_lines:
            cols = [c.strip() for c in line.split('|')[1:-1]]
            if len(cols) == 4:
                step_num = cols[0]
                operation = cols[1]
                test_data = cols[2]
                expected = cols[3]
                steps.append({
                    'step_num': step_num,
                    'operation': operation,
                    'test_data': test_data,
                    'expected': expected
                })

    test_cases.append({
        'case_id': case_id,
        'title': title,
        'test_point': tp,
        'module': module,
        'priority': priority,
        'precondition': precondition,
        'steps': steps
    })

print(f"共解析 {len(test_cases)} 条测试用例")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = '测试用例'

# 定义样式
header_font = Font(bold=True, size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 写入表头
headers = ['用例ID', '用例标题', '关联测试点', '所属模块', '优先级', '前置条件', '步骤编号', '操作步骤', '测试数据', '预期结果']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.alignment = header_alignment
    cell.fill = header_fill
    cell.border = thin_border

# 设置列宽
col_widths = {
    'A': 15,  # 用例ID
    'B': 35,  # 用例标题
    'C': 12,  # 关联测试点
    'D': 15,  # 所属模块
    'E': 8,   # 优先级
    'F': 40,  # 前置条件
    'G': 10,  # 步骤编号
    'H': 40,  # 操作步骤
    'I': 35,  # 测试数据
    'J': 50,  # 预期结果
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# 写入数据
current_row = 2
for case in test_cases:
    # 获取该用例的步骤数
    step_count = len(case['steps'])
    if step_count == 0:
        step_count = 1

    # 合并单元格的范围
    start_row = current_row
    end_row = current_row + step_count - 1

    # 写入用例基本信息（跨行合并）
    case_alignment = Alignment(vertical='center', wrap_text=True)

    ws.cell(row=start_row, column=1, value=case['case_id']).alignment = case_alignment
    ws.cell(row=start_row, column=2, value=case['title']).alignment = case_alignment
    ws.cell(row=start_row, column=3, value=case['test_point']).alignment = case_alignment
    ws.cell(row=start_row, column=4, value=case['module']).alignment = case_alignment
    ws.cell(row=start_row, column=5, value=case['priority']).alignment = case_alignment
    ws.cell(row=start_row, column=6, value=case['precondition']).alignment = case_alignment

    # 合并单元格
    if step_count > 1:
        for col in range(1, 7):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    # 写入步骤
    for i, step in enumerate(case['steps']):
        row = start_row + i
        ws.cell(row=row, column=7, value=step['step_num']).alignment = case_alignment
        ws.cell(row=row, column=8, value=step['operation']).alignment = case_alignment
        ws.cell(row=row, column=9, value=step['test_data']).alignment = case_alignment
        ws.cell(row=row, column=10, value=step['expected']).alignment = case_alignment

    # 添加边框
    for row in range(start_row, end_row + 1):
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border

    current_row = end_row + 1

# 保存文件
output_path = '/home/mxr/e2eProject/testcases_2026-05-13.xlsx'
wb.save(output_path)
print(f"Excel文件已生成: {output_path}")

# 统计模块分布
module_stats = {}
for case in test_cases:
    module = case['module']
    if module:
        module_stats[module] = module_stats.get(module, 0) + 1

print("\n模块分布统计:")
for module, count in sorted(module_stats.items()):
    print(f"  {module}: {count}")
print(f"  总计: {len(test_cases)}")